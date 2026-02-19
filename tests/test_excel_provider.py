import tempfile
import unittest
from pathlib import Path

from ic_tester_app.chips.excel_io import create_excel_template
from ic_tester_app.chips.providers.excel_provider import ExcelChipProvider

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


@unittest.skipIf(load_workbook is None, "openpyxl not installed")
class TestExcelProvider(unittest.TestCase):
    def test_valid_workbook_loads_chip(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "chip_library.xlsx"
            create_excel_template(path, include_sample=True)
            provider = ExcelChipProvider(path)
            chip = provider.get_chip("74181", board="MEGA")
            self.assertIsNotNone(chip)
            self.assertEqual(chip["chipId"], "74181")

    def test_missing_required_column_sets_error(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "chip_library.xlsx"
            create_excel_template(path, include_sample=False)
            wb = load_workbook(path)
            ws = wb["Tests"]
            # Remove expected_json header by replacing with another name.
            ws.cell(row=1, column=6, value="expected_out")
            wb.save(path)

            provider = ExcelChipProvider(path)
            self.assertTrue(provider.has_load_errors())
            self.assertEqual(provider.get_all_chip_ids(board="MEGA"), [])

    def test_malformed_json_sets_error(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "chip_library.xlsx"
            create_excel_template(path, include_sample=True)
            wb = load_workbook(path)
            ws = wb["Tests"]
            ws.cell(row=3, column=5, value='{"A0":"HIGH"')  # broken JSON
            wb.save(path)

            provider = ExcelChipProvider(path)
            self.assertTrue(provider.has_load_errors())
            self.assertIn("Invalid JSON", "\n".join(provider.get_errors()))

    def test_requirement_pullup_pin14_present(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "chip_library.xlsx"
            create_excel_template(path, include_sample=True)
            provider = ExcelChipProvider(path)
            chip = provider.get_chip("74181", board="MEGA")
            self.assertIsNotNone(chip)
            reqs = chip.get("hardwareRequirements", [])
            pullups = [r for r in reqs if r.get("type") == "pullup" and r.get("pin") == 14]
            self.assertTrue(pullups)


if __name__ == "__main__":
    unittest.main()
