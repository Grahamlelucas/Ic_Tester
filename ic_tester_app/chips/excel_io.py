"""
Excel I/O helpers: template generation and results export.
"""

import json
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List

from ..logger import get_logger

logger = get_logger("chips.excel_io")

try:
    from openpyxl import Workbook, load_workbook
except Exception:  # pragma: no cover
    Workbook = None
    load_workbook = None


REQUIRED_SHEETS = {
    "Chips": ["chip_id", "name", "manufacturer", "package", "description", "active"],
    "Pinout": ["chip_id", "pin_number", "pin_name", "direction", "role", "description"],
    "Mappings": ["chip_id", "board", "chip_pin", "arduino_pin", "notes"],
    "Tests": ["chip_id", "test_id", "description", "phase", "input_json", "expected_json", "enabled"],
    "Requirements": ["chip_id", "type", "pin", "signal", "target", "resistor", "description"],
    "Results": [
        "timestamp",
        "chip_id",
        "board",
        "success",
        "tests_run",
        "tests_passed",
        "tests_failed",
        "error",
        "details_json",
    ],
}


def _json_default(value: Any) -> str:
    return json.dumps(value if value is not None else {})


def create_excel_template(path: Path, include_sample: bool = True) -> Path:
    """Create a canonical workbook template."""
    if Workbook is None:
        raise RuntimeError("openpyxl is not installed")

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # Remove default sheet and create required sheets.
    default_sheet = wb.active
    wb.remove(default_sheet)

    for sheet_name, columns in REQUIRED_SHEETS.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(columns)

    if include_sample:
        _append_sample_rows(wb)

    wb.save(path)
    logger.info(f"Created Excel template: {path}")
    return path


def _append_sample_rows(workbook):
    chips_ws = workbook["Chips"]
    pinout_ws = workbook["Pinout"]
    mappings_ws = workbook["Mappings"]
    tests_ws = workbook["Tests"]
    req_ws = workbook["Requirements"]

    chips_ws.append(["74181", "4-Bit Arithmetic Logic Unit (ALU)", "Texas Instruments", "24-pin DIP", "Sample row", True])

    pinout_rows = [
        ["74181", 24, "VCC", "POWER", "VCC", "Power"],
        ["74181", 12, "GND", "POWER", "GND", "Ground"],
        ["74181", 2, "A0", "INPUT", "A0", "Operand A bit 0"],
        ["74181", 1, "B0", "INPUT", "B0", "Operand B bit 0"],
        ["74181", 9, "F0", "OUTPUT", "F0", "Result bit 0"],
        ["74181", 14, "AeqB", "OUTPUT", "AeqB", "Comparator output"],
    ]
    for row in pinout_rows:
        pinout_ws.append(row)

    mapping_rows = [
        ["74181", "MEGA", 24, "PWR_5V", "Power pin"],
        ["74181", "MEGA", 12, "PWR_GND", "Ground pin"],
        ["74181", "MEGA", 1, 22, ""],
        ["74181", "MEGA", 2, 23, ""],
        ["74181", "MEGA", 9, 30, ""],
        ["74181", "MEGA", 14, 34, ""],
    ]
    for row in mapping_rows:
        mappings_ws.append(row)

    tests_ws.append([
        "74181",
        1,
        "Setup defaults",
        "SETUP",
        json.dumps({"A0": "LOW", "B0": "LOW"}),
        "",
        True,
    ])
    tests_ws.append([
        "74181",
        2,
        "Basic test",
        "TEST",
        json.dumps({"A0": "HIGH", "B0": "LOW"}),
        json.dumps({"F0": "HIGH"}),
        True,
    ])

    req_ws.append(["74181", "pullup", 14, "AeqB", "VCC", "10k", "Add a pull-up resistor on pin 14"])


def sync_json_chips_to_excel(
    chips_dir: Path,
    workbook_path: Path,
    board: str = "MEGA",
    preserve_results: bool = True,
) -> Dict[str, int]:
    """
    Build/refresh workbook from JSON chip definitions.
    Writes Chips/Pinout/Mappings/Tests/Requirements and optionally preserves Results.
    """
    if load_workbook is None:
        raise RuntimeError("openpyxl is not installed")

    board_key = str(board or "MEGA").upper()
    chips_dir = Path(chips_dir)
    workbook_path = Path(workbook_path)

    old_results: List[List[Any]] = []
    if preserve_results and workbook_path.exists():
        try:
            old_wb = load_workbook(workbook_path, data_only=True)
            if "Results" in old_wb.sheetnames:
                ws = old_wb["Results"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and any(cell not in (None, "") for cell in row):
                        old_results.append(list(row))
        except Exception:
            old_results = []

    if workbook_path.exists():
        try:
            workbook_path.unlink()
        except Exception:
            pass
    create_excel_template(workbook_path, include_sample=False)
    wb = load_workbook(workbook_path)

    chips_ws = wb["Chips"]
    pinout_ws = wb["Pinout"]
    mappings_ws = wb["Mappings"]
    tests_ws = wb["Tests"]
    req_ws = wb["Requirements"]
    results_ws = wb["Results"]

    json_files = sorted(chips_dir.glob("*.json"))
    chip_rows = 0
    pin_rows = 0
    mapping_rows = 0
    test_rows = 0
    req_rows = 0

    for json_file in json_files:
        with open(json_file, "r") as f:
            chip = json.load(f)

        chip_id = str(chip.get("chipId", "")).strip()
        if not chip_id:
            continue

        chips_ws.append([
            chip_id,
            chip.get("name", chip_id),
            chip.get("manufacturer", ""),
            chip.get("package", ""),
            chip.get("description", ""),
            True,
        ])
        chip_rows += 1

        pinout = chip.get("pinout", {})
        vcc = pinout.get("vcc")
        gnd = pinout.get("gnd")
        if vcc:
            pinout_ws.append([chip_id, int(vcc), "VCC", "POWER", "VCC", "Power"])
            pin_rows += 1
        if gnd:
            pinout_ws.append([chip_id, int(gnd), "GND", "POWER", "GND", "Ground"])
            pin_rows += 1

        for p in pinout.get("inputs", []):
            pinout_ws.append([
                chip_id,
                int(p.get("pin")),
                p.get("name", ""),
                "INPUT",
                p.get("name", ""),
                p.get("description", ""),
            ])
            pin_rows += 1

        for p in pinout.get("outputs", []):
            pinout_ws.append([
                chip_id,
                int(p.get("pin")),
                p.get("name", ""),
                "OUTPUT",
                p.get("name", ""),
                p.get("description", ""),
            ])
            pin_rows += 1

        for nc_pin in pinout.get("noConnect", []):
            pinout_ws.append([chip_id, int(nc_pin), "NC", "NC", "NC", "No connect"])
            pin_rows += 1

        mapping = chip.get("arduinoMapping", {})
        power = mapping.get("power", {})
        io = mapping.get("io", {})

        for chip_pin, target in power.items():
            token = "PWR_5V" if str(target).upper() in {"5V", "VCC"} else "PWR_GND"
            mappings_ws.append([chip_id, board_key, int(chip_pin), token, "Power mapping"])
            mapping_rows += 1

        for chip_pin, ard_pin in io.items():
            mappings_ws.append([chip_id, board_key, int(chip_pin), int(ard_pin), ""])
            mapping_rows += 1

        test_seq = chip.get("testSequence", {})
        for setup in test_seq.get("setup", []):
            step = int(setup.get("step", 1))
            tests_ws.append([
                chip_id,
                step,
                setup.get("action", f"Setup {step}"),
                "SETUP",
                _json_default(setup.get("pins", {})),
                "",
                True,
            ])
            test_rows += 1

        for test in test_seq.get("tests", []):
            test_id = int(test.get("testId", test_rows + 1))
            tests_ws.append([
                chip_id,
                test_id,
                test.get("description", f"Test {test_id}"),
                "TEST",
                _json_default(test.get("inputs", {})),
                _json_default(test.get("expectedOutputs", {})),
                True,
            ])
            test_rows += 1

        for req in chip.get("hardwareRequirements", []):
            if isinstance(req, str):
                req_ws.append([chip_id, "note", "", "", "", "", req])
                req_rows += 1
                continue
            if not isinstance(req, dict):
                continue
            req_ws.append([
                chip_id,
                req.get("type", "note"),
                req.get("pin", ""),
                req.get("signal", ""),
                req.get("target", ""),
                req.get("resistor", ""),
                req.get("description", ""),
            ])
            req_rows += 1

    if preserve_results:
        for row in old_results:
            results_ws.append(row)

    wb.save(workbook_path)
    logger.info(f"Synchronized {chip_rows} JSON chip(s) to workbook {workbook_path}")
    return {
        "chips": chip_rows,
        "pin_rows": pin_rows,
        "mapping_rows": mapping_rows,
        "test_rows": test_rows,
        "requirement_rows": req_rows,
        "results_preserved": len(old_results),
    }


def ensure_required_sheets(workbook_path: Path):
    """Ensure workbook has all required sheets and header row."""
    if load_workbook is None:
        raise RuntimeError("openpyxl is not installed")

    wb = load_workbook(workbook_path)
    updated = False
    for sheet, cols in REQUIRED_SHEETS.items():
        if sheet not in wb.sheetnames:
            ws = wb.create_sheet(sheet)
            ws.append(cols)
            updated = True
        else:
            ws = wb[sheet]
            if ws.max_row < 1:
                ws.append(cols)
                updated = True
    if updated:
        wb.save(workbook_path)
    return wb


def export_results_to_excel(
    results: Dict[str, Any],
    workbook_path: Path,
    board: str = "MEGA",
    sheet_name: str = "Results",
    append: bool = True,
) -> Path:
    """Append a test result row into workbook Results sheet."""
    if load_workbook is None:
        raise RuntimeError("openpyxl is not installed")

    if not workbook_path.exists():
        create_excel_template(workbook_path, include_sample=False)

    wb = ensure_required_sheets(workbook_path)
    ws = wb[sheet_name]

    if not append:
        ws.delete_rows(2, ws.max_row)

    row = [
        datetime.now().isoformat(timespec="seconds"),
        results.get("chipId", ""),
        str(board).upper(),
        bool(results.get("success", False)),
        int(results.get("testsRun", 0)),
        int(results.get("testsPassed", 0)),
        int(results.get("testsFailed", 0)),
        str(results.get("error", "") or ""),
        json.dumps(results.get("testDetails", [])),
    ]
    ws.append(row)
    wb.save(workbook_path)
    logger.info(f"Exported result to workbook: {workbook_path}")
    return workbook_path


def workbook_has_chips(workbook_path: Path, board: str = "MEGA") -> bool:
    """Fast check for at least one active chip in workbook."""
    if load_workbook is None or not workbook_path.exists():
        return False

    wb = load_workbook(workbook_path, data_only=True)
    if "Chips" not in wb.sheetnames or "Mappings" not in wb.sheetnames:
        return False

    chips_ws = wb["Chips"]
    mappings_ws = wb["Mappings"]
    active_ids: List[str] = []

    header = [str(c.value).strip().lower() if c.value is not None else "" for c in chips_ws[1]]
    idx_chip = header.index("chip_id") if "chip_id" in header else -1
    idx_active = header.index("active") if "active" in header else -1
    if idx_chip < 0 or idx_active < 0:
        return False

    for row in chips_ws.iter_rows(min_row=2, values_only=True):
        if not row or idx_chip >= len(row):
            continue
        chip_id = row[idx_chip]
        active = row[idx_active] if idx_active < len(row) else True
        if chip_id and str(active).strip().lower() not in {"false", "0", "no"}:
            active_ids.append(str(chip_id).strip())

    if not active_ids:
        return False

    board_key = str(board).upper()
    mheader = [str(c.value).strip().lower() if c.value is not None else "" for c in mappings_ws[1]]
    try:
        m_chip = mheader.index("chip_id")
        m_board = mheader.index("board")
    except ValueError:
        return False

    for row in mappings_ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        chip_id = row[m_chip] if m_chip < len(row) else None
        mboard = row[m_board] if m_board < len(row) else None
        if chip_id and str(chip_id).strip() in active_ids and str(mboard).strip().upper() == board_key:
            return True
    return False
