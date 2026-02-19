import json
import tempfile
import unittest
from pathlib import Path

from ic_tester_app.chips.database import ChipDatabase
from ic_tester_app.chips.excel_io import create_excel_template

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


class TestChipDatabaseSources(unittest.TestCase):
    def _write_min_json_chip(self, chips_dir: Path, chip_id: str):
        payload = {
            "chipId": chip_id,
            "name": chip_id,
            "pinout": {
                "vcc": 14,
                "gnd": 7,
                "inputs": [{"pin": 1, "name": "A"}],
                "outputs": [{"pin": 2, "name": "Y"}],
            },
            "arduinoMapping": {"power": {"14": "5V", "7": "GND"}, "io": {"1": 22, "2": 23}},
            "testSequence": {
                "tests": [
                    {
                        "testId": 1,
                        "description": "smoke",
                        "inputs": {"A": "HIGH"},
                        "expectedOutputs": {"Y": "HIGH"},
                    }
                ]
            },
        }
        with open(chips_dir / f"{chip_id}.json", "w") as f:
            json.dump(payload, f, indent=2)

    @unittest.skipIf(load_workbook is None, "openpyxl not installed")
    def test_hybrid_unions_excel_and_json(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            chips_dir = root / "chips"
            chips_dir.mkdir()
            self._write_min_json_chip(chips_dir, "7400")

            excel_path = chips_dir / "chip_library.xlsx"
            create_excel_template(excel_path, include_sample=True)  # includes 74181 sample

            db = ChipDatabase(
                chips_dir=chips_dir,
                source_mode="hybrid",
                excel_path=excel_path,
                board="MEGA",
            )
            ids = db.get_all_chip_ids(board="MEGA")
            self.assertIn("7400", ids)
            self.assertIn("74181", ids)


if __name__ == "__main__":
    unittest.main()
