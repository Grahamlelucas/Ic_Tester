"""
Excel-based chip data provider using normalized multi-sheet workbook.
"""

import json
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any

from ...logger import get_logger
from .base import ChipDataProvider

logger = get_logger("chips.providers.excel")

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


class ExcelSchemaError(Exception):
    """Raised when workbook schema or row content is invalid."""


class ExcelChipProvider(ChipDataProvider):
    """
    Loads chip definitions from a normalized Excel workbook:
    Chips, Pinout, Mappings, Tests, Requirements, Results.
    """

    REQUIRED_COLUMNS = {
        "Chips": ["chip_id", "name", "manufacturer", "package", "description", "active"],
        "Pinout": ["chip_id", "pin_number", "pin_name", "direction", "role", "description"],
        "Mappings": ["chip_id", "board", "chip_pin", "arduino_pin", "notes"],
        "Tests": ["chip_id", "test_id", "description", "phase", "input_json", "expected_json", "enabled"],
        "Requirements": ["chip_id", "type", "pin", "signal", "target", "resistor", "description"],
    }

    def __init__(self, workbook_path: Path):
        self.workbook_path = workbook_path
        self._chips_by_board: Dict[str, Dict[str, Dict[str, Any]]] = {"MEGA": {}, "UNO": {}}
        self._errors: List[str] = []
        self.reload()

    def reload(self):
        self._chips_by_board = {"MEGA": {}, "UNO": {}}
        self._errors = []

        if load_workbook is None:
            self._errors.append("openpyxl is not installed")
            logger.error("Excel provider unavailable: openpyxl is not installed")
            return

        if not self.workbook_path.exists():
            self._errors.append(f"Workbook not found: {self.workbook_path}")
            logger.warning(f"Excel workbook not found: {self.workbook_path}")
            return

        try:
            wb = load_workbook(filename=self.workbook_path, data_only=True)
            rows = self._read_and_validate_sheets(wb)
            self._build_chips(rows)
            logger.info(
                f"Excel provider loaded {len(self._chips_by_board['MEGA'])} MEGA chip(s) and "
                f"{len(self._chips_by_board['UNO'])} UNO chip(s)"
            )
        except ExcelSchemaError as e:
            self._errors.append(str(e))
            logger.error(f"Excel schema error: {e}")
        except Exception as e:
            self._errors.append(f"Failed to load workbook: {e}")
            logger.error(f"Failed to load Excel workbook: {e}")

    def _read_and_validate_sheets(self, workbook):
        rows = {}
        for sheet_name, required in self.REQUIRED_COLUMNS.items():
            if sheet_name not in workbook.sheetnames:
                raise ExcelSchemaError(f"Missing required sheet: {sheet_name}")

            ws = workbook[sheet_name]
            header = [self._norm(cell.value) for cell in ws[1]]
            missing = [col for col in required if col not in header]
            if missing:
                raise ExcelSchemaError(
                    f"Sheet '{sheet_name}' missing required columns: {', '.join(missing)}"
                )

            col_index = {name: idx for idx, name in enumerate(header)}
            data = []
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if self._is_blank_row(row):
                    continue
                row_map = {}
                for col_name in required:
                    idx = col_index[col_name]
                    row_map[col_name] = row[idx] if idx < len(row) else None
                row_map["_row"] = row_num
                data.append(row_map)
            rows[sheet_name] = data
        return rows

    @staticmethod
    def _norm(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip().lower()

    @staticmethod
    def _is_blank_row(row: Tuple[Any, ...]) -> bool:
        for cell in row:
            if cell is not None and str(cell).strip() != "":
                return False
        return True

    @staticmethod
    def _as_bool(value: Any, default: bool = True) -> bool:
        if value is None:
            return default
        if isinstance(value, bool):
            return value
        text = str(value).strip().lower()
        if text in {"true", "1", "yes", "y"}:
            return True
        if text in {"false", "0", "no", "n"}:
            return False
        return default

    @staticmethod
    def _as_int(value: Any, field: str, sheet: str, row_num: int) -> int:
        try:
            return int(value)
        except Exception:
            raise ExcelSchemaError(
                f"Invalid integer for {sheet}.{field} at row {row_num}: {value}"
            )

    @staticmethod
    def _parse_json_field(raw: Any, field: str, row_num: int) -> Dict[str, Any]:
        text = "" if raw is None else str(raw).strip()
        if not text:
            return {}
        try:
            parsed = json.loads(text)
        except Exception as e:
            raise ExcelSchemaError(
                f"Invalid JSON in Tests.{field} at row {row_num}: {e}"
            )
        if not isinstance(parsed, dict):
            raise ExcelSchemaError(
                f"Expected JSON object in Tests.{field} at row {row_num}"
            )
        return parsed

    def _build_chips(self, rows: Dict[str, List[Dict[str, Any]]]):
        chips = rows["Chips"]
        pinout_rows = rows["Pinout"]
        mapping_rows = rows["Mappings"]
        test_rows = rows["Tests"]
        req_rows = rows["Requirements"]

        base_chip_defs: Dict[str, Dict[str, Any]] = {}

        for row in chips:
            chip_id = str(row["chip_id"]).strip() if row["chip_id"] is not None else ""
            row_num = row["_row"]
            if not chip_id:
                raise ExcelSchemaError(f"Empty Chips.chip_id at row {row_num}")
            if chip_id in base_chip_defs:
                raise ExcelSchemaError(f"Duplicate chip_id '{chip_id}' in Chips row {row_num}")

            if not self._as_bool(row["active"], default=True):
                continue

            base_chip_defs[chip_id] = {
                "chipId": chip_id,
                "name": str(row["name"] or chip_id),
                "manufacturer": str(row["manufacturer"] or ""),
                "package": str(row["package"] or ""),
                "description": str(row["description"] or ""),
                "pinout": {"inputs": [], "outputs": [], "noConnect": []},
                "arduinoMapping": {"power": {}, "io": {}},
                "testSequence": {"setup": [], "tests": []},
                "hardwareRequirements": [],
            }

        # Pinout
        for row in pinout_rows:
            chip_id = str(row["chip_id"]).strip()
            if chip_id not in base_chip_defs:
                continue
            row_num = row["_row"]
            pin_num = self._as_int(row["pin_number"], "pin_number", "Pinout", row_num)
            pin_name = str(row["pin_name"] or f"PIN{pin_num}").strip()
            direction = str(row["direction"] or "").strip().upper()
            role = str(row["role"] or "").strip().upper()
            desc = str(row["description"] or "").strip()

            chip = base_chip_defs[chip_id]
            pinout = chip["pinout"]

            if direction == "INPUT":
                pinout["inputs"].append({"pin": pin_num, "name": pin_name, "description": desc})
            elif direction == "OUTPUT":
                pinout["outputs"].append({"pin": pin_num, "name": pin_name, "description": desc})
            elif direction == "POWER":
                if role == "VCC":
                    pinout["vcc"] = pin_num
                elif role == "GND":
                    pinout["gnd"] = pin_num
            elif direction == "NC":
                pinout["noConnect"].append(pin_num)
            else:
                raise ExcelSchemaError(
                    f"Invalid Pinout.direction '{direction}' at row {row_num}; "
                    "expected INPUT/OUTPUT/POWER/NC"
                )

        # Tests
        for row in test_rows:
            chip_id = str(row["chip_id"]).strip()
            if chip_id not in base_chip_defs:
                continue
            if not self._as_bool(row["enabled"], default=True):
                continue
            row_num = row["_row"]
            phase = str(row["phase"] or "").strip().upper()
            test_id = self._as_int(row["test_id"], "test_id", "Tests", row_num)
            description = str(row["description"] or f"Test {test_id}").strip()
            inputs = self._parse_json_field(row["input_json"], "input_json", row_num)
            expected = self._parse_json_field(row["expected_json"], "expected_json", row_num)

            chip = base_chip_defs[chip_id]
            if phase == "SETUP":
                chip["testSequence"]["setup"].append(
                    {"step": test_id, "action": description, "pins": inputs}
                )
            elif phase == "TEST":
                chip["testSequence"]["tests"].append(
                    {
                        "testId": test_id,
                        "description": description,
                        "inputs": inputs,
                        "expectedOutputs": expected,
                    }
                )
            else:
                raise ExcelSchemaError(
                    f"Invalid Tests.phase '{phase}' at row {row_num}; expected SETUP/TEST"
                )

        # Requirements
        for row in req_rows:
            chip_id = str(row["chip_id"]).strip()
            if chip_id not in base_chip_defs:
                continue
            req = {
                "type": str(row["type"] or "").strip().lower(),
                "pin": int(row["pin"]) if row["pin"] not in (None, "") else None,
                "signal": str(row["signal"] or "").strip() or None,
                "target": str(row["target"] or "").strip() or None,
                "resistor": str(row["resistor"] or "").strip() or None,
                "description": str(row["description"] or "").strip() or None,
            }
            base_chip_defs[chip_id]["hardwareRequirements"].append(req)

        # Per-board mappings
        for board in ("MEGA", "UNO"):
            board_rows = [
                r for r in mapping_rows
                if str(r["board"] or "").strip().upper() == board
            ]
            mapped_chip_ids = {str(r["chip_id"]).strip() for r in board_rows if r.get("chip_id")}
            per_board_defs = {}
            for chip_id, chip in base_chip_defs.items():
                chip_copy = json.loads(json.dumps(chip))
                chip_copy["arduinoMapping"] = {"power": {}, "io": {}}
                per_board_defs[chip_id] = chip_copy

            for row in board_rows:
                chip_id = str(row["chip_id"]).strip()
                if chip_id not in per_board_defs:
                    continue

                row_num = row["_row"]
                chip_pin = self._as_int(row["chip_pin"], "chip_pin", "Mappings", row_num)
                arduino_pin_raw = row["arduino_pin"]
                if arduino_pin_raw in (None, ""):
                    continue

                pinout = per_board_defs[chip_id]["pinout"]
                mapping = per_board_defs[chip_id]["arduinoMapping"]
                token = str(arduino_pin_raw).strip().upper()

                if token == "PWR_5V":
                    mapping["power"][str(chip_pin)] = "5V"
                elif token == "PWR_GND":
                    mapping["power"][str(chip_pin)] = "GND"
                else:
                    arduino_pin = self._as_int(arduino_pin_raw, "arduino_pin", "Mappings", row_num)
                    if chip_pin == pinout.get("vcc"):
                        mapping["power"][str(chip_pin)] = "5V"
                    elif chip_pin == pinout.get("gnd"):
                        mapping["power"][str(chip_pin)] = "GND"
                    else:
                        mapping["io"][str(chip_pin)] = arduino_pin

            # Validate tests/mapping per chip
            filtered_defs = {}
            for chip_id, chip in per_board_defs.items():
                if not chip["testSequence"]["tests"]:
                    logger.warning(f"Excel chip '{chip_id}' has no TEST rows for board {board}")
                    continue
                if chip_id not in mapped_chip_ids:
                    # No profile for this board by design (e.g. UNO planned later).
                    continue
                if not chip["arduinoMapping"]["io"]:
                    logger.warning(f"Excel chip '{chip_id}' has no I/O mappings for board {board}")
                    continue
                filtered_defs[chip_id] = chip

            self._chips_by_board[board] = filtered_defs

    def get_chip(self, chip_id: str, board: str = "MEGA") -> Optional[Dict[str, Any]]:
        board_key = str(board or "MEGA").upper()
        chips = self._chips_by_board.get(board_key, {})
        return chips.get(chip_id)

    def get_all_chip_ids(self, board: str = "MEGA") -> List[str]:
        board_key = str(board or "MEGA").upper()
        chips = self._chips_by_board.get(board_key, {})
        return sorted(chips.keys())

    def get_chip_count(self, board: str = "MEGA") -> int:
        board_key = str(board or "MEGA").upper()
        return len(self._chips_by_board.get(board_key, {}))

    def validate_chip(self, chip_id: str, board: str = "MEGA") -> Tuple[bool, List[str]]:
        chip = self.get_chip(chip_id, board=board)
        errors = list(self._errors)
        if not chip:
            errors.append(f"Chip '{chip_id}' not found in Excel source")
            return False, errors

        pinout = chip.get("pinout", {})
        if not pinout.get("inputs") and not pinout.get("outputs"):
            errors.append("Pinout must contain inputs and/or outputs")
        if "vcc" not in pinout or "gnd" not in pinout:
            errors.append("Pinout must contain VCC and GND power pins")

        test_seq = chip.get("testSequence", {})
        if not test_seq.get("tests"):
            errors.append("testSequence must contain tests")

        mapping = chip.get("arduinoMapping", {})
        if not mapping.get("io"):
            errors.append(f"No board mapping found for board {str(board).upper()}")

        return len(errors) == 0, errors

    def has_load_errors(self) -> bool:
        return bool(self._errors)

    def get_errors(self) -> List[str]:
        return list(self._errors)
