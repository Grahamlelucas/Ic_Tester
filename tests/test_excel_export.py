import tempfile
import unittest
from pathlib import Path

from ic_tester_app.chips.excel_io import create_excel_template, export_results_to_excel

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


@unittest.skipIf(load_workbook is None, "openpyxl not installed")
class TestExcelExport(unittest.TestCase):
    def test_export_appends_results_row(self):
        with tempfile.TemporaryDirectory() as tmp:
            workbook = Path(tmp) / "chip_library.xlsx"
            create_excel_template(workbook, include_sample=False)

            result = {
                "chipId": "74181",
                "success": True,
                "testsRun": 2,
                "testsPassed": 2,
                "testsFailed": 0,
                "testDetails": [{"testId": 1, "passed": True}],
            }

            export_results_to_excel(result, workbook, board="MEGA")
            wb = load_workbook(workbook, data_only=True)
            ws = wb["Results"]
            self.assertEqual(ws.max_row, 2)
            self.assertEqual(ws.cell(row=2, column=2).value, "74181")
            self.assertEqual(ws.cell(row=2, column=3).value, "MEGA")


if __name__ == "__main__":
    unittest.main()
